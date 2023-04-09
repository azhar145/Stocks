
# C:\Users\aa300j\Downloads\DLP_import_Script_backup5f.py   -- working [without bulk, with 08,03]   run it from streamlit
# Version 1 



def Step2_delta_html_vs_bulk(bulk_folder_path,output_from_DLP_HTML_csv_file_loaded_in_DLP_sucessfully,g3):
    
    import pandas as pd
    import os,numpy
    import sys
    from openpyxl.styles import PatternFill
    from openpyxl.styles import Font
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
    import warnings
    import os
    import openpyxl
    from openpyxl import load_workbook
    from openpyxl.workbook import Workbook
    from openpyxl.styles import Font, Fill
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment
    import shutil
    import streamlit as st


    pd.set_option('display.max_rows', None)
    pd.set_option('display.max_columns', None)
    pd.set_option('display.width', None)


    g3=str(g3)+str('\\')+str('bulk')
    print('g3=  ',g3,'   ', len(os.listdir(g3)))
    print('bulk folder path= ' ,bulk_folder_path,'   ', len(os.listdir(bulk_folder_path)))
    

             
    print('\n')
    print('\n')
    print('--- Input files to current module')
    print('1) ',bulk_folder_path,output_from_DLP_HTML_csv_file_loaded_in_DLP_sucessfully)
    print('2) ',g3)
    print('\n')         


    
    pp=os.listdir(bulk_folder_path)
    
    print('\n\n')
    print('======= bulk folder path / Files === >> [to be loaded in DLP - (with version nos)] << =====', len(pp),' ============ ','\n\n')
    print('\n\n')

    print(g3,' g3')
##    print(pp,' pp')
    print(bulk_folder_path,'   bulk_folder_path')
##    sys.exit()


    

    k=0

    gg=[]

  


    for x in (pp):
        
##        if '*.csv' in x:
##        print(x)
        m=str(x).split('instance~')[1]
        print(m)
        m3=m.split('~')[0]
    ##    print(k,'   ',m3)
        gg.append(m3)
        k=k+1



    print('\n\n')
    print(gg)


    


    print('\n')
    print('======= bulk folder path / Files === >> [to be loaded in DLP - (without version nos)] << =====', len(gg),' ============ ','\n\n')
    print('same as above')
##    print('===== Files already uploaded via bulk-tool in DLP/based on html file exported from DLP ==== Total count=',len(gg))
    print(gg)
    print('\n\n')



    df=pd.read_csv(output_from_DLP_HTML_csv_file_loaded_in_DLP_sucessfully)
    print('\n\n')
    
##    df3=df[df['Created Uid']=='aa300j']
    print(df,'\n')
    df3=df.loc[df['Last Uid']=='aa300j']
    print(df3,' =========== from csv')
    print(df3.shape)

#################
    path = os.path.abspath(__file__)
    dir_path = (path)
    print('\n')
    pp = str('Curr_dir_Curr_Module --> ')+str(dir_path)+str('   ')+str('in module -----> [def Step2_delta_html_vs_bulk]')
    print(pp)
    print('-------------------- filter by [aa300j]')
    print('\n')
################   

    

    p=[]
    for x in df3.index:
##        df3['n'].loc[x]=(df3['Name'].loc[x])
##        df['n'].loc[x]=str(df['Name'].loc[x])+str('~')+str(df['Version'].loc[x])
        p.append(df3['Name'].loc[x])
    ##print(df['n'].sort_values())
##    print('===== output_from_DLP_HTML_csv_file_[loaded_in_DLP successfully]',' Total count=',len(p))    
##    print(p,' 5555     ',len(p))

    print(len(p))
    import pandas as pd

    dt=df.sort_values(by=['Name','Version'], axis=0, ascending=False, inplace=True, kind='quicksort', na_position='last', ignore_index=True, key=None)
    dq=df.drop_duplicates(keep='first')
##    print(dq.columns)


    dt=df[['Name', 'Version']]
    t1=[]
    t2=[]
    for x in dt.index:
        k=1
        for y in dt.index:
            if dt['Name'].loc[x]==dt['Name'].loc[y]:
                k=k+1
                if k==2:
                    t1.append(dt['Name'].loc[x])
                    t2.append(dt['Version'].loc[x])
##                    st.write('Inputs!* :sunglasses:')
                    
                              
                    break

    dm=pd.DataFrame([t1,t2]).T
    print(dm,' highest version')
##    print(dt)


    c=st.container()
    
    
    print('\n')

    nn=len(gg)-len(p)
    print(len(gg),'  ',len(p),'  ',nn)

##    sys.exit()



##    sys.exit()

    import streamlit.components.v1 as components  # Import Streamlit

    # Render the h1 block, contained in a frame of size 200x200.
##    components.html("<html><body>backgroundColor="#f0f0f5";<h1>Hello, World</h1>  </html>", width=200, height=200)
##    st.markdown(f"""<style>.stApp {{background-image: url("https://cdn.pixabay.com/photo/2019/04/24/11/27/flowers-4151900_960_720.jpg");background-attachment: fixed;background-size: cover}}</style>""",unsafe_allow_html=True)           

##    st.markdown(f"""<style>.stApp {{background-image: url("https://cdn.pixabay.com/photo/2019/04/24/11/27/flowers-4151900_960_720.jpg");background-attachment: fixed;background-size: cover}}</style>""",unsafe_allow_html=True)           
##    st.markdown(f"""<style>.stApp {{backgroundColor = '#00325B}}</style>""",unsafe_allow_html=True)
    st.markdown("""<style>.reportview-container {background: url("https://images.app.goo.gl/LFCobouKtT7oZ7Qv7")}</style>""",unsafe_allow_html=True)
    
    k=1
    print('In bulk [delta b/w [Candidate vs Producion]','   ',len(gg),'   ',len(set(gg)))

    x33='In bulk'
##    st.markdown(f'<h3 style="color:#333dff;font-size:24px;">{x33}</h3>', unsafe_allow_html=True)

    st.markdown(f'<p style="color:#333dff; font-size: 36px;">{x33}</h3>', unsafe_allow_html=True)
        
        


    for x in (gg):
        print(k,'   ',x)
        st.markdown(f'<p small style="color:#333dff;font-size:10px;margin:0;padding:0;line-height:0px;">{k,x}</small></p>', unsafe_allow_html=True)
        k=k+1

    print('\n\n\n')
    k=1
    print('in html/DLP server','   ',len(p),'     ',len(set(p)))

    x34='in html/DLP server'
    st.markdown(f'<h3 style="color:#333dff;font-size:24px;">{x34}</h3>', unsafe_allow_html=True)

    
    p2=p.sort()
    for x in (p):
        print(k,'   ',x)
        st.markdown(f'<p small style="color:#333dff;font-size:10px;margin:0;padding:0;line-height:0px;">{k,x}</small></p>', unsafe_allow_html=True)
        k=k+1


    single_counts = set()
    duplicates = []

    for x in p:
        if x in single_counts:
            duplicates.append(x)
        else:
            single_counts.add(x)

    print('\n\n')
    print('to check, if any duplicates in dlp/html: ')
    print('single counts= ',len(single_counts),'\n',single_counts)
    print('\n')
    print('duplicates count= ',len(duplicates),'\n',duplicates)
    print('\n')

##
##    print(gg,' 0000000')
##    print('\n\n')

##    print(p,' 88888888')
    temp3 = list(set(gg) - set(p))

    ##print(temp3,'  delta      ',len(temp3))

    print('==== Missing ones are ========== ',len(temp3))
    import subprocess
    k=1

    print(' ===================   Missing files to be uploaded manually   ============== ')
    print('\n')

    ######### need to fix this (copy to a folder not working)

    
    for x in temp3:
    ##    print(str(bulk_folder_path))
        print(k,') ',x)
        k=k+1
##        tt=str('copy')+str(' ')+str(bulk_folder_path)+str('\\')+str(str('instance~')+str(x)+str('.csv')).strip()+ str(' ')+str(g3)
##        print(tt)
##        pu = os.popen(tt).read()
##        print('5555 ',pu)


    print('\n\n')
    print(len(temp3),' files - need to enter manually')
    print('\n\n')
    
    mm=r'C:\Users\aa300j\Downloads\nn2'
    mm5=r'C:\Users\aa300j\Downloads\nn5'

    for x in os.listdir(mm):
        cv=os.path.join(mm,x)
        os.remove(cv)
        print(cv)


    print('\n\n')
    for x in os.listdir(mm5):
        cv=os.path.join(mm5,x)
        os.remove(cv)
        print(cv)    
            
    print('\n\n')
    for x in os.listdir(str(r'C:\Users\aa300j\Downloads\Candidate_DLP_files_x4')):
        path55=os.path.join(str(r'C:\Users\aa300j\Downloads\Candidate_DLP_files_x4'),str(x))
        for y in temp3:
            if str(y) in str(x):
##                print(x,' 999999999999999')
                shutil.copy(path55, str(mm)+str('\\')+str(x))
    print('\n\n')
    
    for x in os.listdir(str(r'C:\Users\aa300j\Downloads\Candidate_DLP_files_x4\bulk')):
        path55=os.path.join(str(r'C:\Users\aa300j\Downloads\Candidate_DLP_files_x4\bulk'),str(x))
        for y in temp3:
            if str(y) in str(x):
##                print(x,' 999999999999999')
                shutil.copy(path55, str(mm5)+str('\\')+str(x))


    print('\n\n')
    print('Missing files are: (non bulk) ',str(mm),'  ',len(os.listdir(mm)))
    print('\n\n')
    print('Missing files are (bulk): ',str(mm5),'   ',len(os.listdir(mm5)))
    print(' ============================================= rrr ================================================')












    sys.exit() 
######## tttt azhar delete below

    
    mm=r'C:\Users\aa300j\Downloads\nn2'
    k=1
    for x in (pp):
    
    ##        if '*.csv' in x:
    ##        print(x)
        m=str(x).split('instance~')[1]
##        print(m)
        m3=m.split('~')[0]
##        print(k,'   ',m3)
        for kp in temp3:
            if m3==kp:
                path55=os.path.join(str(bulk_folder_path),str(x))
                
                path56=os.path.join(str(mm),str(x))
##                if not os.path.exists(path55):
##                    print("Path55 of the file is Invalid")
##                if not os.path.exists(path56):
##                    print("Path56 of the file is Invalid")    
                
                print(k,' --------- found ---------------- ',m3,'  ',bulk_folder_path)
                shutil.copy(path55, str(mm)+str('\\')+str(x))
                
                k=k+1 

    print(len(temp3),' missing files - need to enter manually are located in -----> ',mm)
    print(dm,' highest version')

            
     

    path = os.path.abspath(__file__)
    dir_path = (path)
    print('\n')
    pp = str('Curr_dir_Curr_Module --> ')+str(dir_path)+str('   ')+str('in module -----> [def Step2_delta_html_vs_bulk]')
    print('\n')
    print('##################################################')
    print('***** where am i?   ')
    print(pp)
    print('filter by [aa300j]')
    print('******************** 334')
    print('##################################################')
    print('\n')
    print('\n')

################   



    sys.exit() 

##    p87=r'C:\Users\aa300j\Downloads\Candidate_DLP_files_x4\bulk'
##    if os.path.isfile(p87)==True:
##        shutil.rmtree(p87)
##    os.mkdir(r'C:\Users\aa300j\Downloads\Candidate_DLP_files_x4\bulk3')            

     

        

    print('\n\n')
    print('============== Summary ====================')
    print('Missing files are located at [should be uploaded to DLP] ',g3)
    print('# of files in bulk ',len(gg),'  bulk')
    print('# of files in html [uniques] ',len(set(p)),'   html')
    print('# of files to be uploaded manually ',len(temp3),'   ','delta')
    print('\n')
    print(dm,' highest version')
    print('\n')
    print(bulk_folder_path,'  ',pp, ' bulk folder path code 443 ')

##    print('# of files in DLP: ',len(p))




###########################################################################################################################################################################
###########################################################################################################################################################################
###########################################################################################################################################################################
###########################################################################################################################################################################
###########################################################################################################################################################################


        
def Step1_compare_Production_Candidate__delete_same_ones(Candidate_DLP_files_folder,Production_DLP_files_folder,g3,g4,g5,g3a):
    import pandas as pd
    import os,numpy
    import sys
    from openpyxl.styles import PatternFill
    from openpyxl.styles import Font
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
    import warnings
    import os
    import openpyxl
    from openpyxl import load_workbook
    from openpyxl.workbook import Workbook
    from openpyxl.styles import Font, Fill
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment
    from openpyxl import load_workbook
    from openpyxl.workbook import Workbook
    import os.path
    import os
    import subprocess
    import shutil
    import fnmatch
    import streamlit as st

    pd.set_option('display.max_rows', None)
    pd.set_option('display.max_columns', None)
    pd.set_option('display.width', None)

#######################
    path = os.path.abspath(__file__)
    dir_path = (path)
    print('\n')
    pp = str('Curr_dir_Curr_Module --> ')+str(dir_path)+str('   ')+str('in module -----> [Step1_compare_Production_Candidate__delete_same_ones]')
    print('***** where am i?   ')
    print(pp)
    print('******************** 334')
    print('\n')
################


    
    p87=r'C:\Users\aa300j\Downloads\Candidate_DLP_files_x4\bulk'
    p88=r'C:\Users\aa300j\Downloads\Candidate_DLP_files_x4'

##    for f in p87:
##        print(f)



    print(os.listdir())

##    if os.path.isfile(p87)==True:
    if os.path.exists(p87):
        shutil.rmtree(p87)
    if os.path.exists(p88):
        shutil.rmtree(p88)

##    p87=r'C:\Users\aa300j\Downloads\Candidate_DLP_files_x4'
####    if os.path.isfile(p87)==True:
##    shutil.rmtree(p87)
##    
##    os.mkdir(r'C:\Users\aa300j\Downloads\Candidate_DLP_files_x4')
##        

##    os.mkdir(r'C:\Users\aa300j\Downloads\Candidate_DLP_files_x4')

        
    if os.path.exists(p88)==False:
        os.mkdir(r'C:\Users\aa300j\Downloads\Candidate_DLP_files_x4')
    if os.path.exists(p87)==False:
        os.mkdir(r'C:\Users\aa300j\Downloads\Candidate_DLP_files_x4\bulk')

    g3=r'C:\Users\aa300j\Downloads\Candidate_DLP_files_x4'



    

    if not (os.listdir(str(g3))):
        pass
##        print('hhhhhhhhhhhhhhhhhhhhhhhh')
    else:
        k=0
        for f in os.listdir(str(g3)):
            path=os.path.join(g3, f )

##            print(path)
            try:
                os.remove(path)

##                print(k,'  ',path,'  removed')
                k=k+1
            except OSError as e: # name the Exception `e`
                print('\n')
                print("             Failed with:", e.strerror) # look what it says
##                print("Error code:", e.code)
                exc_type, exc_obj, exc_tb = sys.exc_info()
                fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                print(              exc_type, fname, exc_tb.tb_lineno)
                print('\n')
            

#    removing files from bulk folder:
    for f in os.listdir(str(g3b)):
        path=os.path.join(g3b,f)
        os.remove(path)


    if not (os.listdir(str(g4))):
        pass
##        print('hhhhhhhhhhhhhhhhhhhhhhhh')
    else:
        k=0
        for f in os.listdir(str(g4)):
            path=os.path.join(g4, f )
##            print(path)
            try:
                os.remove(path)
##                print(k,'  ',path,'  removed')
                k=k+1
            except OSError as e: # name the Exception `e`
                print('\n')
                print("             Failed with:", e.strerror) # look what it says
##                print("Error code:", e.code)
                exc_type, exc_obj, exc_tb = sys.exc_info()
                fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                print(              exc_type, fname, exc_tb.tb_lineno)
                print('\n')

    print(g3,'   ')        
##    sys.exit()
    
    
    for x in os.listdir(Candidate_DLP_files_folder):

        if '.csv' in x:

            path1=os.path.join(Candidate_DLP_files_folder,x)
            path2=os.path.join(g3,x)

    ##        print('Candidate   ',path1)
    ##        print('\n\n')
    ##        print('Production ',path2)
            
    ##        print(str(Candidate_DLP_files_folder)+str('/')+str(x))
    ##        shutil.copy(str(Candidate_DLP_files_folder)+str('\\')+str(x), str(g3)+str('\\')+str(x))
            shutil.copy(path1,path2)
        

##    print('\n\n\n')
    print(path1,'  path1')
    print(path2,'  path2')
    

    
    print('\n\n')
##    print(Production_DLP_files_folder)
##    print(g4)
##    sys.exit()

##    sys.exit()
    for x2 in os.listdir(Production_DLP_files_folder):

        if '.csv' in x2:
            path3=os.path.join(Production_DLP_files_folder,x2)
            path4=os.path.join(g4,x2)
    ##        print(path1)
    ##        print(str(Production_DLP_files_folder)+str('/')+str(x))
    ##        shutil.copy(str(Production_DLP_files_folder)+str('\\')+str(x2), str(g4)+str('\\')+str(x2))    
            shutil.copy(path3,path4)

    print(Production_DLP_files_folder,'  path3')
    print(g4,'  g4')
##    print(path4,'  path4')

##    sys.exit()   
##    sys.exit() 

##    bulk_folder_path=r'C:\Users\aa300j\Downloads\PEP\Connect_it_files\Data_Loading\Data Management\create_DLP_files\DLP_NCX_template_(open In-Service file)\C2s_December_1_2022\HST5G\Candidate'
##    output_from_DLP_HTML_csv_file_loaded_in_DLP_sucessfully=r'C:\Users\aa300j\Desktop\Data Dictionary and Lookup Platform_HST5g_uploaded.csv'

####################





############
############    for x in os.listdir(Candidate_DLP_files_folder):
##############        print(x)
##############        u=str('copy')+str('  ')+ str(Candidate_DLP_files_folder)+str('\\')+str(x)+str('  ')+str(g3)
############        u=str('ls ')+str(str(Candidate_DLP_files_folder)+str('\\')+str(x))
##############        print(u)
############        t=os.popen(u)
############        print(t.read(),' *************** ')
############        
############
############    sys.exit()    
        
##    src_dir = Candidate_DLP_files_folder     
##    # path to destination directory
##    dest_dir = g3
##     
##    # getting all the files in the source directory
##    files = os.listdir(src_dir)
##     
##    shutil.copytree(src_dir, dest_dir)
###################
##    src_dir = Production_DLP_files_folder     
##    # path to destination directory
##    dest_dir = g4
##     
##    # getting all the files in the source directory
##    files = os.listdir(src_dir)
##     
##    shutil.copytree(src_dir, dest_dir)
##################

    g3_count = len(fnmatch.filter(os.listdir(g3), '*.*'))
    g4_count = len(fnmatch.filter(os.listdir(g4), '*.*'))

    pp55_g3=[]
    for x in os.listdir(g3):
        if '.csv' in x:
            pp55_g3.append(x)

    pp55_g4=[]
    for x in os.listdir(g4):
        if '.csv' in x:
            pp55_g4.append(x)
        



    pp=os.listdir(g3)
    pp2=os.listdir(g4)
    print('Candidate [duplicate deleted/Should be loaded in DLP]   ',len(pp),'\n',str(pp))
    print('\n\n')
    print('Production [from previous DLP load from shared drive] ',len(pp2),'\n',str(pp2))
    k=0
    import csv
    gg=[]
    print('\n\n')
    tt=[]
##    z5=['vlanCharacteristicsInstanceProdRegion','ProdZone']
    z5=['mooooooooon']
    k=1
    s=1

##    print(g3,' ------------ g3 ----------------')
##    print('\n')
##    print(pp2,' --------- pp ----------')
##    sys.exit()

    print('\n\n')
    print('g3= ',g3)
    print('g4= ',g4)

    print('\n\n')
    



    Deleted_files_Common=[]
    Deleted_files_dir=[]
    to_be_loaded_in_DLP_bulk=[]
    
    dfd=pd.DataFrame()
    for x in (pp):
        
        
        if x not in z5:
        

            for y in (pp2):

                if '.csv' in str(x) and '.csv' in str(y):

##                if '~' in str(x) and '~' in str(y) and '.csv' in str(x) and '.csv' in str(y):

                    x2=str(x).split(' ')[1]
                    y2=str(y).split(' ')[1]
##                    print(x2,' 44444444444444444444444444444444444444444444444444444444444444444  ')
##                    print(y2,' 555555555555555555555555555555555555555555555555555555555555555555  ')
##
##                    print('\n\n\n')
                    


##                    print(str(x).split('~')[1],' fffffffffffff')
##                    print(str(y).split('~')[1],' ttttttttttttt')

##                sys.exit()
                    if str(x2)==str(y2):
                        
        ##                print(x)
                        df = pd.read_csv(str(g3)+str('\\')+str(x))
        ##                print(' ========== ',k,'    ',x)
                        print(df,' --------- > ',x)

                        df2 = pd.read_csv(str(g4)+str('\\')+str(y))
        ##                print(' ========== ',k,'    ',y)
                        print(df2,' ------> ',y)

                        df5=pd.concat([df,df2],axis=0)
                        df5.drop_duplicates(keep=False, inplace=True,ignore_index=False)
                        
                        if df.equals(df2):
                            pp5=df.equals(df2)
    ##                        print('\n')
    ##                        print('Common files in Candidate and Production [to - be - removed]','\n',pp5)
                            os.remove(str(g3)+str('\\')+str(x))
                            
    ##                        tt=str('DELETE ')+ str(g3)+str('\\')+str(x)
    ##                        print('        ',tt)
    ##                        p = os.popen(tt).read()
    ##                        print('\n')
    ##                        print('         ',p)
    ##                        print('Common files Deleted Candidate_DLP_files_x4/Candite folder',g3,'   ',x)
                            Deleted_files_Common.append(x)
                            Deleted_files_dir.append(g3)

                            
                            
                            
    ##                        pass
    ##                    else:
    ##                        print(s,')     ',x)
    ##                        s=s+1
    ##
    ##
            
    ##
    ##                       
    ##
    ##                        file_exists = os.path.exists(str(g3)+str('\\')+str(x))
    ##                        if file_exists == True:
    ##                           print('dddddddddddddddddddddddddddddddddddddddddddd')
    ##
    ##                        
    ##    ##                    tt=str('copy ')+ str(g3)+str('\\')+str(x) + str('  ')+str(g5)
    ##    ##                    tt=str('copy ')+ str(r'C:\\Users\\aa300j\\Downloads\\Candidate_DLP_files_x4')+str('\\')+str(x) + str('  ')+str('C:\\Users\\aa300j\\Downloads\\Delta_Cand_vs_Product_[to_go_to_DLP_after_bulk_change]_x6')
    ##                        tt=str("copy ")+(str(g3)+str('\\')+str(x))+str('  ')+str(g5)
    ##                        
    ##    ##                ##    os.system(t)
    ##    ##                ##    os.system("copy C://Users//aa300j//Downloads//p34//instance_relationshipTenantSIInstanceProdZoneATN3_2.6.csv C://Users//aa300j//Downloads//p34//mm.csv")
    ##    ##                    tt = os.popen(str('copy C:\\Users\\aa300j\\Downloads\\p34\\instance_relationshipTenantSIInstanceProdZoneATN3_2.6.csv C:\\Users\\aa300j\\Downloads\\p34\\m')).read() 
    ##    ##                ##    subprocess.run(t)
    ##    ##                    print(tt)
    ##    ##                    p = os.popen(tt).read()
    ##    ##                    print(p, ' to ', str(g5)+str('\\'))
    ##
    ##                        print('        ',tt)
    ##                        p = os.popen(tt).read()
    ##                        print('\n')
    ##                        print('         ',p)
    ##                                    
                        


                k=k+1

    print('\n\n')

    print('here')

    
##    print('Common Files deleted back into Candidate_DLP_files_x4/Candidate folder')
##    print('# of Candidate DLP files [Common files deleted from Candidate] = ',len(pp))
##    print('# of Production DLP files = ',len(pp2))
##    print('\n')
##    print('what files to be loaded in DLP based on delta ====  ',r'C:\Users\aa300j\Downloads\Candidate_DLP_files_x4')
##    print('\n')
##    print('to_be_loaded_in_dlp_bulk 33 df5.shape ',df5.shape,' 00000000000000 df5')
##    print('to_be_loaded_in_dlp_bulk',len(to_be_loaded_in_DLP_bulk))
##    print('azhar no_of_files_to_be_uploaded :',count)

    count = len(fnmatch.filter(os.listdir(g3), '*.*')) 
##    print('# of files in Candidate folder: ',g3_count,'         ',g3)
##    print('# of files in Production folder: ',g4_count,'        ',g4)
##    print('# of files in Candidate folder: ',len(pp55_g3),'  33')
##    print('# of files in Production folder: ',len(pp55_g4),'   33')
    import os

    print('\n\n')
    print('# of files in Candidate folder: ',len(pp55_g3),'      ',Candidate_DLP_files_folder,'    code 214')
    print('# of files in Production folder: ',len(pp55_g4),'      ',Production_DLP_files_folder,'    code 214')
    

    candx=[]
    for x in os.listdir(g3):
        if '.csv' in x:
            candx.append(candx)
            
    prodx=[]
    for y in os.listdir(g4):
        if '.csv' in y:
            prodx.append(prodx)




##        for y in list(g4):
##            if str(x)==str(y):
##                break
##            elif len(y)==k-1 and str(x)!=str(y):
##                print(str(x),end=' ')
##
##            k=k+1


    bbulk=(str(g3)+str('\\')+str('bulk'))
    print(bbulk)
##    print(os.listdir(bbulk),' 8888cc')
    

    count=len(candx)

    print('# of DLP files to be uploaded to bulk',count,'        ',g3,'  code 214')
    print('# of DLP files to be uploaded to bulk',(os.listdir(bbulk)),'        ',bbulk,'  code 214')
    print('\n')
    print('Bulk too -----')
##    st.text('=================')
##    st.write('g3 --> ',g3)
##    st.write('bbulk ---> ',bbulk)
##    for x in prodx:
##        st.text(prodx)
#######################
    path = os.path.abspath(__file__)
    dir_path = (path)
    print('\n')
    pp = str('Curr_dir_Curr_Module --> ')+str(dir_path)+str('   ')+str('in module -----> [Step1_compare_Production_Candidate__delete_same_ones]')
    print('***** where am i?   ')
    print(pp)
    print('******************** 334')
    print('\n')

################ 

    depend_DLP_import_script.change_filenames_for_bulk_import(version_no,g3)




###########################################################################################################################################################################
###########################################################################################################################################################################
###########################################################################################################################################################################
###########################################################################################################################################################################
###########################################################################################################################################################################


def run_step1(Candidate_DLP_files_folder,Production_DLP_files_folder,g3,g4,g5,g3b):
    ######### Start of Step 1 #######################################################################

    import os

    st.write('Inputs!* :sunglasses:')
    st.markdown(f'<p small style="color:#333dff;font-size:10px;margin:0;padding:0;line-height:0px;">{Candidate_DLP_files_folder,len(os.listdir(Candidate_DLP_files_folder))}</small>', unsafe_allow_html=True)
    st.markdown(f'<p small style="color:#333dff;font-size:10px;margin:0;padding:0;line-height:0px;">{Production_DLP_files_folder,len(os.listdir(Production_DLP_files_folder))}</small>', unsafe_allow_html=True)

##    st.write('Candidate folder = ', Candidate_DLP_files_folder)
##    st.write('Production folder = ', Production_DLP_files_folder)
    st.write(":heavy_minus_sign:" * 34) # horizontal separator line.

    st.write('Temporary!* :sunglasses:')
    st.markdown(f'<p small style="color:#333dff;font-size:10px;margin:0;padding:0;line-height:0px;">{g3,len(os.listdir(g3))}</small>', unsafe_allow_html=True)
    st.markdown(f'<p small style="color:#333dff;font-size:10px;margin:0;padding:0;line-height:0px;">{g4,len(os.listdir(g4))}</small>', unsafe_allow_html=True)
    st.write(":heavy_minus_sign:" * 34) # horizontal separator line.
    hh='List of files to be imported to bulk-DLP'
    with st.container():
            
        st.markdown(f'<h3 style="color:#bb33ff;font-size:14px;">{hh}</h3>', unsafe_allow_html=True)
    ##    st.markdown(f'<h3 style="color:#bb33ff;font-size:14px;">{g3b}</h3>', unsafe_allow_html=True)


        
        
        Step1_compare_Production_Candidate__delete_same_ones(Candidate_DLP_files_folder,Production_DLP_files_folder,g3,g4,g5,g3b)
        print('\n')
        print('==============   Summary =====================')
        print('\n')
        print('<----------------> def Step1_compare_Production_Candidate__delete_same_ones(Candidate_DLP_files_folder,Production_DLP_files_folder,g3,g4,g5,g3b)',' <------------->')
        print('--- input files:')
        print('Input Canidate folder : ',Candidate_DLP_files_folder,'    ',len(os.listdir(Candidate_DLP_files_folder)),' files','  ',os.listdir(Candidate_DLP_files_folder)[-1])
        print('Input Production folder : ',Production_DLP_files_folder,'    ',len(os.listdir(Production_DLP_files_folder)),' files','  ',os.listdir(Production_DLP_files_folder)[-1])
        print('\n\n')
        print(' --- output files [to go into DLP] ------ > ')
        print(' Delta Diff files that should be difference between Prod vs candidate are located in : code 332',)
        print('\n')
        print(' =======================   ',g3,' ==========================  ')
        print('\n')
        print(' # of files :',len(os.listdir(g3))-1,'  files afer removing duplicates b/w production and candidate and removed vlanchar and zones')
        ##print(os.listdir(g3),' 999992222')
        ##print(os.listdir(g3)[-2])
        print(os.listdir(g3))
        print('\n\n\n')


        z5=['vlanCharacteristicsInstanceProdRegion','ProdZone']
        for x in (g3):
            
        ##        if 'vlanCharacteristicsInstanceProdRegion' in x or 'ProdZone' in x:
        ##            print(x)
            k=0
            if str(z5) in str(x):
                print(x)
                k=k+1

        if k==0:
            print('None of 3 files found -----> ',z5,'   in ',g3)



    ##    Step1_compare_Production_Candidate__delete_same_ones(Candidate_DLP_files_folder,Production_DLP_files_folder,g3,g4,g5,g3a)    
        ##        
        ####################################################################################################### End of step 1
        ##    tt=str('copy')+str(' ')+str(g3)+str('\\')+str(x) + str(' ')+str(g4)
        ##    print(tt)
        ##    p = os.popen(tt).read()
        ##    print(p)
            

def run_step2(bulk_folder_path,output_from_DLP_HTML_csv_file_loaded_in_DLP_sucessfully,g3):
    import streamlit as st
    ################################################# Step 2   ##########################
    print('\n\n\n')
    print('*********************************************************************************************************************************************************')
    print('*********************************************************************************************************************************************************')
    print('*********************************************************************************************************************************************************')
    print('*********************************************************************************************************************************************************')
    print('Start of step2 ')
    print('*********************************************************************************************************************************************************')
    print('*********************************************************************************************************************************************************')
    print('*********************************************************************************************************************************************************')
    print('*********************************************************************************************************************************************************')
    print('\n\n\n')
    print(bulk_folder_path)
    print(output_from_DLP_HTML_csv_file_loaded_in_DLP_sucessfully)
    print(g3)
    print('\n\n\n')
    
        
    Step2_delta_html_vs_bulk(bulk_folder_path,output_from_DLP_HTML_csv_file_loaded_in_DLP_sucessfully,g3)



########################################### End of step 2 ##########################
def callback(version_no,Candidate_DLP_files_folder,Production_DLP_files_folder,output_from_DLP_HTML_csv_file_loaded_in_DLP_sucessfully):
   
    import streamlit as st
    st.write('started processing ==========')

    st.markdown("<span style='color:red'>Update/upgrade</span>",
                 unsafe_allow_html=True)


    run_step1(Candidate_DLP_files_folder,Production_DLP_files_folder,g3,g4,g5,g3a)
    run_step2(bulk_folder_path,output_from_DLP_HTML_csv_file_loaded_in_DLP_sucessfully,g3)



############################################# End of step 1 ##########################

##remove_3_files(g3)


###################################################################
import streamlit as st
#####################################################################################################################
#### input files
##version_no = st.text_input('Enter DLP version no: ')
##version_no=2.8


#####################################################################################################################

###LSA4E  # for 1st def
Candidate_DLP_files_folder =r'C:\Users\aa300j\Downloads\PEP\Connect_it_files\bingo_3332\Cand'
Production_DLP_files_folder=r'C:\Users\aa300j\Downloads\PEP\Connect_it_files\bingo_3332\Prod'

##Candidate_DLP_files_folder=st.text_input('Candidate folder (with 008) ')
##Production_DLP_files_folder=st.text_input('Production folder (with 008) ')


output_from_DLP_HTML_csv_file_loaded_in_DLP_sucessfully=str(r'C:\Users\aa300j\Desktop\MPL1a_33333.csv')
                                                            ## Do not change these

##text_input = st.text_area("Enter a text", key="input_text", on_change=callback,args=(text_input,Candidate_DLP_files_folder,Production_DLP_files_folder,output_from_DLP_HTML_csv_file_loaded_in_DLP_sucessfully))

import os
import depend_DLP_import_script
import streamlit as st

print('Candidate folder ',Candidate_DLP_files_folder,' # of files in ',len(os.listdir(Candidate_DLP_files_folder)))
print('Production folder ',Production_DLP_files_folder,' # of files in ',len(os.listdir(Candidate_DLP_files_folder)))
print('\n\n')


## Do not change these
g3=r'C:\Users\aa300j\Downloads\Candidate_DLP_files_x4'      ## Do not change these
g3b=r'C:\Users\aa300j\Downloads\Candidate_DLP_files_x4\bulk' 
# Candidate
g4=r'C:\Users\aa300j\Downloads\Production_DLP_files_x5'     ## Do not change these 
# Production
bulk_folder_path=r'C:\Users\aa300j\Downloads\Candidate_DLP_files_x4\bulk'   ## Do not change these
#### output files
g5='p'
##g5=r'C:\Users\aa300j\Downloads\Delta_Cand_vs_Product_[to_go_to_DLP_after_bulk_change]_x6'   # Delta DLPs to be uploaded after running bulk upload.

g3a=''


##print('\n')
##print('Delta diff files b/w html and Delta Diff, code 333')
##remove_3_files()

#change_filenames_for_bulk_import(version_no,g3)


##st.text_input(on_change=callback(string,Candidate_DLP_files_folder,Production_DLP_files_folder,output_from_DLP_HTML_csv_file_loaded_in_DLP_sucessfully), key='text_key')
##
##
##

with st.form(key="form1",clear_on_submit=True):
    
    version_no = st.text_input(label="DLP_Version_no")
##    output_from_DLP_HTML_csv_file_loaded_in_DLP_sucessfully=st.text_input(label='MPLA.csv')
    
##    submit=st.form_submit_button_button(label='submit')
    submit=st.form_submit_button()
    st.write(version_no)

    if "load_state" not in st.session_state:
         st.session_state.load_state = True
    
    if submit:

        callback(version_no,Candidate_DLP_files_folder,Production_DLP_files_folder,output_from_DLP_HTML_csv_file_loaded_in_DLP_sucessfully)


##title = st.text_input("label goes here")
##version_no=title
##
##if title:
####    df = pd.read_csv(path)
####    st.experimental_rerun()
##    
##    callback(float(version_no),Candidate_DLP_files_folder,Production_DLP_files_folder,output_from_DLP_HTML_csv_file_loaded_in_DLP_sucessfully)
##else:
##    if not title:
##        st.warning("Please fill out so required fields")
